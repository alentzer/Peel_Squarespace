import openpyxl
from openpyxl.styles import PatternFill

# -----------------------------
# FILE PATHS
# -----------------------------
MAPPING_FILE = "/workspaces/Peel_Squarespace/tag_adder/validation_list.xlsx"
PRODUCT_FILE = "/workspaces/Peel_Squarespace/tag_adder/squarespace_current.xlsx"
OUTPUT_FILE = "/workspaces/Peel_Squarespace/tag_adder/squarespace_current_UPDATED.xlsx"

# -----------------------------
# LOAD WORKBOOKS
# -----------------------------
mapping_wb = openpyxl.load_workbook(MAPPING_FILE)
mapping_ws = mapping_wb.active   # first sheet

product_wb = openpyxl.load_workbook(PRODUCT_FILE)
product_ws = product_wb.active   # first sheet

# -----------------------------
# BUILD CATEGORY → TAGS LOOKUP
# -----------------------------
category_to_tags = {}

for row in mapping_ws.iter_rows(min_row=2, values_only=True):
    category, tags = row
    if category:
        category_to_tags[str(category).strip()] = str(tags).strip() if tags else ""

# -----------------------------
# YELLOW HIGHLIGHT STYLE
# -----------------------------
yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# -----------------------------
# COLUMN INDEXES (1-based)
# -----------------------------
CATEGORY_COL = 27   # AA
TAGS_COL = 28       # AB

# -----------------------------
# PROCESS PRODUCTS
# -----------------------------
for row in range(2, product_ws.max_row + 1):

    category_cell = product_ws.cell(row=row, column=CATEGORY_COL)
    tags_cell = product_ws.cell(row=row, column=TAGS_COL)

    category_value = category_cell.value

    # Skip missing categories
    if not category_value:
        continue

    category_value = str(category_value).strip()

    # Strict match
    if category_value not in category_to_tags:
        continue

    new_tags = category_to_tags[category_value]

    # Replace tags entirely
    tags_cell.value = new_tags
    tags_cell.fill = yellow

# -----------------------------
# SAVE UPDATED FILE
# -----------------------------
product_wb.save(OUTPUT_FILE)
print("Done. Updated file saved as:", OUTPUT_FILE)
